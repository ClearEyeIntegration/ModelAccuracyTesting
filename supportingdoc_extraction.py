from datetime import date
from datetime import datetime
import pandas as pd
import random
from dateutil import parser
import time
import requests
from fuzzywuzzy import fuzz
from openpyxl.reader.excel import load_workbook
from util import get_access_token, list_jobs,list_NT_lc_txns, assign_lc_to_user,product_type, user_id,user_id_rob,username_rob, environment, excel_sheet,\
    upload_api, Excel_Path, accuracy_exclude_list, redFill, bill_type_api, GOODS_API,org_code,branch,UPLOAD_WAIT,doc_extraction_api
#from requests_html import HTMLSession
#import runner

counter_ref = 0
#wait_time = 0


def upload_files(authorization_key):
    
    # Method to upload files
    #return: The reference ID list
    
    #global _token
    global txn_created_count
    global prod_type
    global ref_id
    global extracted_file_name #document file name -extracted from API response
    global document_type_id 
    
    
    #global ext_file_name
    #_token = token
    #print(_token)
    try:
      
        #global wait_time
        reference_id_list = []
        txn_created_count = 0
        #document_type_id = excel_sheet.cell(row=2, column=7).value
        #rnum = random.random()
        #wait_time = excel_sheet.cell(row=2, column=8).value
        #wait_time = excel_sheet.cell(row=2, column=8).value
        
        for each_row in range(2, excel_sheet.max_row + 1):
            file_name = excel_sheet.cell(row=each_row, column=2).value
            file_path = excel_sheet.cell(row=each_row, column=3).value
            execute_status = excel_sheet.cell(row=each_row, column=4).value
            #wait_time = excel_sheet.cell(row=each_row, column=8).value
            #wait_time = excel_sheet.cell(row=2, column=8).value
            #tester = excel_sheet.cell(row=each_row, column=7).value
            #amitabh commented below
            reference_id = excel_sheet.cell(row=each_row, column=5).value
            
            c_time = datetime.today().strftime("%I:%M:%S %p")
            #reference_id = excel_sheet.cell(row=each_row, column=2).value + c_time
          
            #form_type = excel_sheet.cell(row=each_row, column=6).value
            if execute_status == 'Yes':
                files = {'file': open(file_path + "\\" + file_name , 'rb')}
                #reference_id = tester + excel_sheet.cell(row=each_row, column=2).value + c_time
                #file_name_concat = file_name[:-4]      #removing last 4characters
                #reference_id = tester + file_name_concat + c_time
                       
                today = date.today()
  
                presentation_date = c_time
               
                headers001 = {
                    'organization_code': org_code,
                    'project_type': 'trade-finance',
                    'id': 'SecRef_' + str(reference_id),
                    'is_enquiry_mode':'false',
                    #'origin':'https://ce-tf-ui-qa-v2.psionix.dev',
                    'preferredusername':'maker',
                    'product_type':'',
                    'presentation_date': presentation_date,
                    'upload_product_type':product_type,
                    #'template_name':'',
                    #'transaction_reference':'',
                    'upload_type':'1',
                    'branch':branch,
                    'presentation_sequence': str(reference_id),   #RefID
                     'Authorization': authorization_key,
                   
                    'UserName':'Rob Maker',
                    'userid':'7fc7b4ca-d58d-4d94-b4d6-86fc05757980',
                    'role': 'maker',
                    'Accept': 'application/json, text/plain, */*' 
                    }
                               

                data = {
                         "field_meta": "{}",
                         
                        }
        
            
                upload_response = requests.post(upload_api,data=data,files=files, headers=headers001)
                time.sleep(UPLOAD_WAIT)    #amitabh added
                #time.sleep(wait_time)    #amitabh added
                print(upload_response.json()) #Sabin added for testing

                if upload_response.status_code == 202:
                    print(file_name + " is uploaded successfully to " + reference_id)
                    if reference_id not in reference_id_list:
                        reference_id_list.append(reference_id)
                        txn_created_count +=1
                else:
                    print(upload_response.status_code)
            else:
                print("execute_status is 'No' in Upload List")

        print(reference_id_list)
        return reference_id_list

    except Exception as error:
        print("Error: couldn't upload the file "+ str(error))

#list_of_NewTask_LC_Txns
def list_of_jobs(authorization_key, reference_id_list):
    document_type_id = excel_sheet.cell(row=2, column=7).value
    header002 = {
        'organization_code': org_code,
        'Authorization': authorization_key,
        'service_type_id': '1',
        'status': '1',
        'is_enquiry_mode': 'false',
        'preferredusername': 'maker',
        'role': 'maker',
        'product_type': '',
        'transaction_reference':'',
        'userid':user_id_rob ,
        'username':username_rob

    }

    
   
    payload = {
        "view": 'new',
        "tag": 'Letter of credit',
        "page_size": 50,      #will fetch additional 10txns-same env someone else creating txn
        "page_no": 1
    }

    header_columns = ["Sl No", "File Name", "Document Type", "Tag Name", "Actual Values", "Confidence Score"]
    header_df = pd.DataFrame(columns=header_columns)
    with pd.ExcelWriter(Excel_Path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as wb:
        header_df.to_excel(wb, sheet_name='Extraction_Sheet', index=False)

        
        try:
           
            list_jobs_response = requests.get(list_NT_lc_txns, headers=header002, json=payload)
            print(list_jobs_response.json())
            if list_jobs_response.status_code == 200 or list_jobs_response.status_code == 201:
                assign_wid(list_jobs_response, authorization_key,reference_id_list)
                #get_bill_types(list_jobs_response, authorization_key,reference_id_list)
            else:
                print('job list api failed')
                
        except Exception as e:
            print('Exception'  +str(e))
        #get_bill_types(authorization_key,reference_id_list,document_type_id)  #doc types : coverletter,BOE,BOL
    total_ref = len(reference_id_list)
    #global counter_ref
    #for each_ID in reference_id_list:
        #counter_ref += 1
        #if counter_ref <= total_ref:
    get_bill_types(authorization_key,reference_id_list,document_type_id)  #doc types : coverletter,BOE,BOL
        #else:        
            #print("All Yes-Txnsaction References extracted ")


def assign_wid(list_jobs_response, authorization_key,reference_id_list):
    
   
     
    jobs_list = list_jobs_response.json()['result']    #inside result-> txns(work_item_ids) are available
    print(jobs_list)                 #pulled from dashboard first/2nd page response
    
    for each_ID in reference_id_list:    #reference_id_list - ref list from excel 2nd column
        #wid_num = 1
        for item in jobs_list:
            if each_ID == item["reference"]:
                print('match')
                prod_type = item["product_type"]
                ref_id = str(each_ID)
                for item2 in item["work_items"]:#first item[work_items] will bring docv id, so break after first id,else it will bring 4ids-docv/doce/tbml
                    #assign_response=assign_lc_to_user(item2["id"], each_ID, authorization_key,item["product_type"])
                    #assign_response=assign_lc_to_user(item2["id"], each_ID, authorization_key,prod_type,ref_id,str(item2["id"])
                    assign_lc_to_user(item2["id"], each_ID, authorization_key,prod_type,ref_id,str(item2["id"]))
              

def get_bill_types(authorization_key,reference_id_list,document_type_id):
    
    
   
    #prod_type = ''
    #ref_id = ''
    #global wait_time
    #jobs_list = list_jobs_response.json()['result']    #inside result-> txns(work_item_ids) are available
    #print(jobs_list)
    total_ref = len(reference_id_list)
    global counter_ref
    for each_ID in reference_id_list:
        counter_ref += 1
        if counter_ref <= total_ref:
        
            ref_id = str(each_ID)
               # print('Match found-Txn created in CT for refid:' + ref_id)
                #prod_type = item["product_type"]
                
               
            header003 = {
                        'organization_code': org_code,
                        'Authorization': authorization_key,
                        'Is_enquiry_mode': 'true',
                        'preferredusername': 'maker',
                        'role': 'maker',
                        #'product_type': item["product_type"],
                        'product_type': 'ELCP',
                        'userid':user_id_rob ,
                        'username':username_rob,
                        'Lc_id':'',
                        'Transaction_reference':ref_id
                        }
            time.sleep(UPLOAD_WAIT)
            document_types_response = requests.get(bill_type_api, headers=header003)
            print(document_types_response.json())
            if document_types_response.status_code == 200:
                    read_extraction_documents(document_types_response,document_type_id, ref_id, authorization_key)
            else:
                    print("bill_types api response is not 200 for the reference ID "+ ref_id)
        else:        
            print("All Yes-Txnsaction References extracted ")
          
            

    

        #else:
           # print("Unable to find reference ID, If it is already assigned, you wont be able to see it here")
    #txn_num += 1

def goods_data_extraction(goods_response,document_file_name):
    response_json = goods_response.json()

    final_df = pd.DataFrame()
    extraction_data = response_json['result']['groups'][0]['goods_description']
    tag_names, values, confidence_score, file_name, bill_type = ([] for _ in range(5))
    file_name = document_file_name
    bill_type = response_json['result']['groups'][0]['document_group_name']
    for each_val in extraction_data:
        gds_desc = (each_val['description']['value'])
        gds_hscode = (each_val['hs_code']['value'])
        gds_qnty = (each_val['quantity']['value'])
        gds_unit = (each_val['unit']['value'])
        gds_amount = (each_val['amount']['value'])
        gds_amount_curency = (each_val['amount_currency']['value'])
        gds_unitprice = (each_val['unit_price']['value'])
        gds_unitcurrency = (each_val['unit_price_currency']['value'])

        # to get confidence value
        gds_desc_conf = 0
        gds_hscode_conf = 0
        gds_qnty_conf = 0
        gds_unit_conf = 0
        gds_amount_conf = 0
        gds_amount_currency_conf = 0
        gds_unitprice_conf = 0
        gds_unitcurrency_conf = 0
        net_weight_conf = 0
        net_weight_unit_conf = 0
        gross_weight_conf = 0
        gross_weight_unit_conf = 0

        confidence_score.append(gds_desc_conf)
        confidence_score.append(gds_hscode_conf)
        confidence_score.append(gds_qnty_conf)
        confidence_score.append(gds_unit_conf)
        confidence_score.append(gds_amount_conf)
        confidence_score.append(gds_amount_currency_conf)
        confidence_score.append(gds_unitprice_conf)
        confidence_score.append(gds_unitcurrency_conf)
        confidence_score.append(net_weight_conf)
        confidence_score.append(net_weight_unit_conf)
        confidence_score.append(gross_weight_conf)
        confidence_score.append(gross_weight_unit_conf)

        tag_names.append('Goods Description')
        values.append(gds_desc)
        tag_names.append('gds_hscode')
        values.append(gds_hscode)
        tag_names.append('gds_qnty')
        values.append(gds_qnty)
        tag_names.append('gds_unit')
        values.append(gds_unit)
        tag_names.append('gds_amount')
        values.append(gds_amount)

        tag_names.append('gds_amount_curency')
        values.append(gds_amount_curency)
        tag_names.append('gds_unitprice')
        values.append(gds_unitprice)
        tag_names.append('gds_unitcurrency')
        values.append(gds_unitcurrency)

        net_weight = (each_val['net_weight']['value'])
        net_weight_unit = (each_val['net_weight_unit']['value'])
        gross_weight = (each_val['gross_weight']['value'])
        gross_weight_unit = (each_val['gross_weight_unit']['value'])

        tag_names.append('gds_net_weight')
        values.append(net_weight)
        tag_names.append('gds_net_weight_unit')
        values.append(net_weight_unit)
        tag_names.append('gds_gross_weight')
        values.append(gross_weight)
        tag_names.append('gds_gross_weight_unit')
        values.append(gross_weight_unit)

    df = pd.DataFrame.from_dict(
        {'Sl No': '', 'File Name': file_name, 'Document Type': bill_type,
         'Tag Name': tag_names, 'Actual Values': values, 'Confidence Score': confidence_score})
    final_df = pd.concat([final_df, df])
  
    with pd.ExcelWriter(Excel_Path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as wb:
        #print(wb.sheets['Extraction_Sheet'].max_row)
        final_df.to_excel(wb, sheet_name='Extraction_Sheet', header=False, index=False,
                          startrow=wb.sheets['Extraction_Sheet'].max_row)
    print("Extraction of all the documents (Goods) is written to the Excel sheet")


def get_goods_services(doc_type_id,ref_id,authorization_key,document_file_name):

    goods_url = GOODS_API
   
    payload005 = {}
    headers005 = {
                'authority': 'ce-tf-api-qa.psionix.dev',
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'en-US,en;q=0.9',
                #'authorization': 'eyJhbGciOiJSUzI1NiIsInR5cCIgOiAiSldUIiwia2lkIiA6ICJjLXcxTzRXQ0ZXTG9RWFJtbFpkMlY5eEk4X1hDSEhsVWRwYkNyRFU1NjBZIn0.eyJleHAiOjE2OTgzNDU4NTcsImlhdCI6MTY5ODMwOTg1OSwiYXV0aF90aW1lIjoxNjk4MzA5ODU3LCJqdGkiOiJlMmE5ZmE3My1mZGQwLTQ3YTAtYWQ2Ny0zNTVlNDFjOTYzMzMiLCJpc3MiOiJodHRwczovL2tleWNsb2FrLnBzaW9uaXguZGV2L2F1dGgvcmVhbG1zL1RyYWRlRmluYW5jZSIsImF1ZCI6ImFjY291bnQiLCJzdWIiOiI3ZmM3YjRjYS1kNThkLTRkOTQtYjRkNi04NmZjMDU3NTc5ODAiLCJ0eXAiOiJCZWFyZXIiLCJhenAiOiJ0cmFkZWZpbmFuY2VfZnJvbnRlbmQiLCJub25jZSI6IjgxZTQ3NzgwLTQ4NjAtNDM0ZS04ZjNhLTg4OWVlNzk3NGY1ZiIsInNlc3Npb25fc3RhdGUiOiJjMGYzMzE0My02ZDI3LTQ1NzgtODhmMy0yNGQ0ZmIyZGIzNWUiLCJhY3IiOiIxIiwiYWxsb3dlZC1vcmlnaW5zIjpbImh0dHBzOi8vY2UtdGYtdWktZGV2LnBzaW9uaXguZGV2IiwiaHR0cDovL2NlLW5vdW4tZXh0cmFjdGlvbi11aS1kZXYubG9jYWxob3N0OjQyMTAiLCJodHRwczovL3N0ZWFzdHVzbWZwb2MuejEzLndlYi5jb3JlLndpbmRvd3MubmV0IiwiaHR0cDovL2dpYi1jZS10Zi11aS1xYS12Mi5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vY29tcGxpYW5jZS1zdGFnZS5jbGVhcnRyYWRlLmRldiIsImh0dHBzOi8vZ2liLWNlLXRmLXVpLXFhLXYyLnBzaW9uaXguZGV2IiwiaHR0cDovL2xvY2FsaG9zdDo0MjAwIiwiaHR0cDovL3dvcmtiZW5jaC1zdGFnZS12Mi5jbGVhcnRyYWRlLmRldi5sb2NhbGhvc3Q6NDIwMCIsImh0dHA6Ly9heGlzLWNlLXRmLXVpLWRldi12Mi5sb2NhbGhvc3Q6NDIwMCIsImh0dHA6Ly9naWItY2UtdGYtdWktZGV2LXYyLmxvY2FsaG9zdDo0MjAwIiwiaHR0cHM6Ly9jdC1jb21wbGlhbmNlLXVpLXFhLmNsZWFydHJhZGUuZGV2IiwiaHR0cDovL2NlLXRmLXVpLWRldi5sb2NhbGhvc3Q6NDIxMCIsImh0dHA6Ly9jZS1ub3VuLWV4dHJhY3Rpb24tdWktZGV2LmxvY2FsaG9zdDo0MjAwIiwiaHR0cDovL2NlLXNoZWxsLXVpLWRldi5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vY2UtdGYtdWktcWEtdjIucHNpb25peC5kZXYiLCJodHRwOi8vMTA0LjIxMS4yMTIuOTU6ODA4NCIsImh0dHA6Ly9jb21wbGlhbmNlLWRldi5jbGVhcnRyYWRlLmRldiIsImh0dHBzOi8vY2xlYXJ0cmFkZS5wc2lvbml4LmRldiIsImh0dHA6Ly9jZS10Zi11aS1xYS5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vanBtYy5jbGVhcnRyYWRlLmRldiIsImh0dHA6Ly9jb21wbGlhbmNlLWRldi5jbGVhcnRyYWRlLmRldi5sb2NhbGhvc3Q6NDIyMCIsImh0dHBzOi8vY2UtdGYtdWktZGV2LXYyLnBzaW9uaXguZGV2IiwiaHR0cDovL3dvcmtiZW5jaC1zdGFnZS5jbGVhcnRyYWRlLmRldi5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vY2UtdGYtdXBsb2Fkcy1xYS5wc2lvbml4LmRldiIsImh0dHBzOi8vY2UtdGYtdWktcWEucHNpb25peC5kZXYiLCJodHRwczovL3dvcmtiZW5jaC1zdGFnZS12Mi5jbGVhcnRyYWRlLmRldiIsImh0dHA6Ly9qcG1jLWNlLXRmLXVpLWRldi5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vd29ya2JlbmNoLXN0YWdlLmNsZWFydHJhZGUuZGV2IiwiaHR0cDovL2NlLXRmLXVpLWRldi5sb2NhbGhvc3Q6NDIwMCIsImh0dHBzOi8vY29tZXJpY2EtY2UtdGYtdWktZGV2LmxvY2FsaG9zdDo0MjAwIl0sInJlYWxtX2FjY2VzcyI6eyJyb2xlcyI6WyJtYWtlciJdfSwicmVzb3VyY2VfYWNjZXNzIjp7ImFjY291bnQiOnsicm9sZXMiOlsibWFuYWdlLWFjY291bnQiLCJtYW5hZ2UtYWNjb3VudC1saW5rcyIsInZpZXctcHJvZmlsZSJdfX0sInNjb3BlIjoib3BlbmlkIHByb2ZpbGUgZW1haWwiLCJzaWQiOiJjMGYzMzE0My02ZDI3LTQ1NzgtODhmMy0yNGQ0ZmIyZGIzNWUiLCJlbWFpbF92ZXJpZmllZCI6ZmFsc2UsInNzb19lbmFibGVkIjpmYWxzZSwibmFtZSI6IlJvYiBNYWtlciIsInByZWZlcnJlZF91c2VybmFtZSI6Im1ha2VyIiwiZ2l2ZW5fbmFtZSI6IlJvYiIsImZhbWlseV9uYW1lIjoiTWFrZXIiLCJvcmdhbml6YXRpb25fY29kZSI6IkNFVFIiLCJlbWFpbCI6Im1ha2VyQGNsZWFyZXllLmFpIiwiYnJhbmNoX3JvbGVfcHJvZHVjdCI6WyJIS19NQUtFUl9FTENQIl19.PnhpM1Rm0-zn6qzJJZZaNmH8dERsSXEdTiZqY3CPLYJBZgW_B_Kyw443DX_Cm-fFtFgR1pklx-fKSaHtDk8L6sc2AWIq-WG5gPLaThXoE9jGTRGYucq8cxs9PndTYuy3DgNWJMnzNxOQjIss5_vCGO1ElnSGaOfsn3kSoTPCISFKOmXYCjyJOW76kS2OFQHx2FPJ4rqS3x2xb5ZBTQOLSIrxUl46RnRPjDVl00EfBUITxVIRb75qAfJl3b475uViJC36bPl405-TdWIdPQF2QQ_zyNDBlBkXVRkONGHBBmEw_pfXFGVKRdC_J8RXkvmDCKSs6IgfWRvjVqRNxK9G9Q',
                'authorization': authorization_key,

                #'document_type_id': '4',
                'document_type_id': str(doc_type_id),
                'is_enquiry_mode': 'true',
                'lc_id': '',
                'organization_code': org_code,
                #'origin': 'https://ce-tf-ui-qa-v2.psionix.dev',
                'preferredusername': 'maker',
                'product_type': 'ELCP',
                #'referer': 'https://ce-tf-ui-qa-v2.psionix.dev/',
                'role': 'maker',
                'sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-site',
                #'transaction_reference': 'TESTQANS_4100',
                'transaction_reference': ref_id,
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
                'userid': user_id_rob,
                'username': username_rob
            }

    goods_response  = requests.request("GET", GOODS_API, headers=headers005, data=payload005)

    print(goods_response.text)

    if goods_response.status_code == 200:
        goods_data_extraction(goods_response,document_file_name)
    else:
        print("Goods services status code is not 200")
    #print(goods_response)


def supporting_document_extracted_data(read_extraction_response):
    response_json = read_extraction_response.json()

    final_df = pd.DataFrame()
    extraction_data = response_json['result'][0]['extract'][0]['extracts']
    tag_names, values, confidence_score, file_name, bill_type = ([] for _ in range(5))
    extracted_file_name = (response_json['result'][0]['document_file_name'])
    bill_type = (response_json['result'][0]['document_type_name'])
    #ext_file_name = extracted_file_name

    for key, value in extraction_data.items():
        if key != 'metadata':
            for each_val in value:
                key_name = (each_val['metadata']['key'])
                key_conf = (each_val['metadata']['confidence'])
                rule_type = (each_val['metadata']['rule_type'])
                if each_val['metadata']['rule_type'] == 'name_address':
                    key_value_name = (each_val['name']['value'])
                    # Need to enable at the time of production push

                    key_value_country = (each_val['country']['value'])

                    # to get confidence value
                    key_value_name_conf = (each_val['name']['metadata']['confidence'])
                    key_value_country_conf = (each_val['country']['metadata']['confidence'])
                    confidence_score.append(key_value_name_conf)
                    confidence_score.append(key_value_country_conf)
                    confidence_score.append('NA')

                    tag_names.append(key_name + '_name'.replace('_', ' '))
                    values.append(key_value_name)
                    try:
                        key_value_address = (each_val['address']['value'])
                    except:
                        key_value_address = ''
                    key_value_referencenumber = (each_val['reference_number']['value'])
                    key_value_identifiercode = (each_val['identifier_code']['value'])
                    tag_names.append(key_name + '_address'.replace('_', ' '))
                    values.append(key_value_address)
                    tag_names.append(key_name + '_country'.replace('_', ' '))
                    values.append(key_value_country)
                    tag_names.append(key_name + '_referencenumber'.replace('_', ' '))
                    values.append(key_value_referencenumber)
                    tag_names.append(key_name + '_identifiercode'.replace('_', ' '))
                    values.append(key_value_identifiercode)

                elif each_val['metadata']['rule_type'] == 'goods_description':
                    gds_desc = (each_val['description']['value'])
                    gds_hscode = (each_val['hs_code']['value'])
                    gds_qnty = (each_val['quantity']['value'])
                    gds_unit = (each_val['unit']['value'])
                    gds_amount = (each_val['amount']['value'])
                    gds_amount_curency = (each_val['amount_currency']['value'])
                    gds_unitprice = (each_val['unit_price']['value'])
                    gds_unitcurrency = (each_val['unit_price_currency']['value'])

                    # to get confidence value
                    gds_desc_conf = (each_val['description']['metadata']['confidence'])
                    gds_hscode_conf = (each_val['hs_code']['metadata']['confidence'])
                    gds_qnty_conf = (each_val['quantity']['metadata']['confidence'])
                    gds_unit_conf = (each_val['unit']['metadata']['confidence'])
                    gds_amount_conf = (each_val['amount']['metadata']['confidence'])
                    gds_amount_currency_conf = (each_val['amount_currency']['metadata']['confidence'])
                    gds_unitprice_conf = (each_val['unit_price']['metadata']['confidence'])
                    gds_unitcurrency_conf = (each_val['unit_price_currency']['metadata']['confidence'])
                    confidence_score.append(gds_desc_conf)
                    confidence_score.append(gds_hscode_conf)
                    confidence_score.append(gds_qnty_conf)
                    confidence_score.append(gds_unit_conf)
                    confidence_score.append(gds_amount_conf)
                    confidence_score.append(gds_amount_currency_conf)
                    confidence_score.append(gds_unitprice_conf)
                    confidence_score.append(gds_unitcurrency_conf)

                    tag_names.append('Goods Description')
                    values.append(gds_desc)
                    tag_names.append(key_name + '_gds_hscode'.replace('_', ' '))
                    values.append(gds_hscode)
                    tag_names.append(key_name + '_gds_qnty'.replace('_', ' '))
                    values.append(gds_qnty)
                    tag_names.append(key_name + '_gds_unit'.replace('_', ' '))
                    values.append(gds_unit)
                    tag_names.append(key_name + '_gds_amount'.replace('_', ' '))
                    values.append(gds_amount)

                    tag_names.append(key_name + '_gds_amount_curency'.replace('_', ' '))
                    values.append(gds_amount_curency)
                    tag_names.append(key_name + '_gds_unitprice'.replace('_', ' '))
                    values.append(gds_unitprice)
                    tag_names.append(key_name + '_gds_unitcurrency'.replace('_', ' '))
                    values.append(gds_unitcurrency)

                elif each_val['metadata']['rule_type'] == 'quantity':
                    quantity_vol = (each_val['volume']['value'])
                    quantity_unit = (each_val['unit']['value'])


                    # to get confidence value
                    quantity_vol_conf = (each_val['volume']['metadata']['confidence'])
                    quantity_unit_conf = (each_val['unit']['metadata']['confidence'])
                    confidence_score.append(quantity_vol_conf)
                    confidence_score.append(quantity_unit_conf)

                    tag_names.append((key_name + '_quantity_vol').replace('_', ' '))
                    values.append(quantity_vol)
                    tag_names.append((key_name + '_quantity_unit').replace('_', ' '))
                    values.append(quantity_unit)
                    # tag_names.append((key_name + '_quantity_amount').replace('_', ' '))
                    # values.append(quantity_amount)

                elif each_val['metadata']['rule_type'] == 'seaport':
                    seaport_name = (each_val['name']['value'])
                    seaport_country = (each_val['country']['value'])
                    seaport_intended = (each_val['is_intended']['value'])

                    # to get confidence value
                    seaport_name_conf = (each_val['name']['metadata']['confidence'])
                    seaport_country_conf = (each_val['country']['metadata']['confidence'])
                    seaport_intended_conf = (each_val['is_intended']['metadata']['confidence'])
                    confidence_score.append(seaport_name_conf)
                    confidence_score.append(seaport_country_conf)
                    confidence_score.append(seaport_intended_conf)

                    tag_names.append((key_name + '_name').replace('_', ' '))
                    values.append(seaport_name)
                    tag_names.append((key_name + '_country').replace('_', ' '))
                    values.append(seaport_country)
                    tag_names.append((key_name + '_intended').replace('_', ' '))
                    values.append(seaport_intended)


                elif each_val['metadata']['rule_type'] == 'airport':
                    airport_name = (each_val['name']['value'])
                    airport_country = (each_val['country']['value'])
                    airport_iatacode = (each_val['iata_code']['value'])

                    # to get confidence value
                    airport_name_conf = (each_val['name']['metadata']['confidence'])
                    airport_country_conf = (each_val['country']['metadata']['confidence'])
                    airport_iatacode_conf = (each_val['iata_code']['metadata']['confidence'])
                    confidence_score.append(airport_name_conf )
                    confidence_score.append(airport_country_conf)
                    confidence_score.append(airport_iatacode_conf)

                    tag_names.append((key_name + '_name').replace('_', ' '))
                    values.append(airport_name)
                    tag_names.append((key_name + '_country').replace('_', ' '))
                    values.append(airport_country)
                    tag_names.append((key_name + '_intended').replace('_', ' '))
                    values.append(airport_iatacode)

                elif each_val['metadata']['rule_type'] == 'language':
                    lan_name = (each_val['name']['value'])
                    lan_code = (each_val['code']['value'])

                    # to get confidence value
                    lan_name_conf = (each_val['name']['metadata']['confidence'])
                    lan_code_conf = (each_val['code']['metadata']['confidence'])
                    confidence_score.append(lan_name_conf)
                    confidence_score.append(lan_code_conf)

                    tag_names.append((key_name + '_lan_name').replace('_', ' '))
                    values.append(lan_name)
                    tag_names.append('Language Of Document')
                    values.append(lan_code)

                elif each_val['metadata']['rule_type'] == 'amount':
                    if each_val['metadata']['key'] != 'amount_in_words':
                        amount = (each_val['total']['value'])
                        currency = (each_val['currency']['value'])

                        # to get confidence value
                        key_value_amount = (each_val['total']['metadata']['confidence'])
                        key_value_currency = (each_val['currency']['metadata']['confidence'])
                        confidence_score.append(key_value_amount)
                        confidence_score.append(key_value_currency)

                        tag_names.append('Amount')
                        values.append(amount)
                        tag_names.append('Currency')
                        values.append(currency)
                    else:
                        # New change
                        amount_words = (each_val['total_in_words']['value'])
                        currency_words = (each_val['currency_in_words']['value'])

                        # to get confidence value
                        key_value_amount_words = (each_val['total_in_words']['metadata']['confidence'])
                        key_value_currency_words = (each_val['currency_in_words']['metadata']['confidence'])
                        confidence_score.append(key_value_amount_words)
                        confidence_score.append(key_value_currency_words)

                        tag_names.append('Amount in Words')
                        values.append(amount_words)
                        tag_names.append('Currency from Amount in words')
                        values.append(currency_words)

                elif each_val['metadata']['rule_type'] == 'vessel':
                    name = (each_val['name']['value'])

                    # to get confidence score
                    vessel_name_conf = (each_val['name']['metadata']['confidence'])
                    # vessel_is_intended_conf = (each_val['is_intended']['metadata']['confidence'])
                    confidence_score.append(vessel_name_conf)
                    # confidence_score.append(vessel_is_intended_conf)

                    # newly added the condition for checking the is intended
                    if each_val['is_intended'] != "None":
                        is_intended = (each_val['is_intended']['value'])
                        tag_names.append(key_name + 'is_intended')
                        values.append(is_intended)
                        confidence_score.append('NA')

                    tag_names.append('Vessel')
                    values.append(name)
                    # tag_names.append('Is Intended')
                    # values.append(is_intended)

                # else condition satisfies the rule type : string, tenor, place, boolean
                else:
                    key_value = (each_val['value'])
                    key_conf_value = (each_val['metadata']['confidence'])
                    tag_names.append(key_name.replace('_', ' '))
                    values.append(key_value)
                    confidence_score.append(key_conf_value)
        #important module
        df = pd.DataFrame.from_dict(
                {'Sl No': '', 'File Name': extracted_file_name, 'Document Type': bill_type,
                 'Tag Name': tag_names, 'Actual Values': values, 'Confidence Score': confidence_score})
    final_df = pd.concat([final_df, df])

    #print("Extraction of all the documents is written to the Excel sheet")
    with pd.ExcelWriter(Excel_Path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as wb:

        #print(wb.sheets['Extraction_Sheet'].max_row)
        final_df.to_excel(wb, sheet_name='Extraction_Sheet', header=False, index=False,
                          startrow=wb.sheets['Extraction_Sheet'].max_row)
    
    print("Extraction_DataVerification_Fields of " + str(extracted_file_name) + "completed and data written to the Excel sheet")
    return extracted_file_name


# to extract the support document
'''def read_extraction_documents(document_types_response,ref_id, authorization_key):
  
    supporting_documents = document_types_response.json()['result']
    #print(type(supporting_documents))
    print(str(supporting_documents)) #amitabh added

    for each in supporting_documents:
        if each['name'].lower != 'unclassified':
            if each['received_original_count']:
                try:
                    # read extraction
                    #read_extraction = bill_type_api + str(lc_ids) + '/extraction/document'
                    read_extraction = doc_extraction_api
                    #doc_type_id = str(each['document_type_id'])
                    header004 = {
                        'organization_code': org_code,
                        'document_type_id': str(each['document_type_id']), #id extracted from classifier response
                        #'document_type_id':doc_type_id, #id extracted from classifier response
                        'Authorization': authorization_key,
                        'userid': user_id,
                        'Is_enquiry_mode': 'false',
                        'preferredusername': 'maker',
                        'role': 'maker',
                        'product_type': 'ELCP',
                        'userid':user_id_rob ,
                        'username':username_rob,
                        'Lc_id':'',
                        'Transaction_reference':str(ref_id),
                        'Accept':'application/json, text/plain, */*',
                        'Origin':'https://ce-tf-ui-qa-v2.psionix.dev',
                        'Referer':'https://ce-tf-ui-qa-v2.psionix.dev/'
                    }
                    #time.sleep(60)
                    read_extraction_response = requests.get(read_extraction, headers=header004,timeout=30)
                    print(str(read_extraction_response))

                 
                    if read_extraction_response.status_code == 200:
                        # Extracting supporting document field values
                        print('supporting document extraction started for '+each['name'])
                        #supporting_document_extracted_data(read_extraction_response)
                        extracted_file_name = supporting_document_extracted_data(read_extraction_response)
    
                    else:
                        print("Failed to extract supporting documents "+ each['name'])
                    try:
                     get_goods_services(each['document_type_id'],ref_id,authorization_key,extracted_file_name)
                    except Exception as e:
                      print('goods service api failed for '+each['document_file_name'] + ' with exception '+str(e))
                except Exception as e:
                    print("Failed to extract supporting documents " + each['name'] +'and the exception is '+str(e))
'''
# to extract the support document
def read_extraction_documents(document_types_response,document_type_id,ref_id, authorization_key):
    #global wait_time
    supporting_documents = document_types_response.json()['result']
    

    for each in supporting_documents:
        #if each['name'].lower != 'unclassified':
        #if (each['name'].lower != 'unclassified') and (each['document_type_id'] == document_type_id):
        if each['document_type_id'] == document_type_id:
            if each['received_original_count']:
                try:
                    # read extraction                    #read_extraction = bill_type_api + str(lc_ids) + '/extraction/document'
                    read_extraction = doc_extraction_api
                    #doc_type_id = str(each['document_type_id'])
                    header004 = {
                        'organization_code': org_code,
                        'document_type_id': str(each['document_type_id']), #id extracted from classifier response
                        #'document_type_id':doc_type_id, #id extracted from classifier response
                        'Authorization': authorization_key,
                        'userid': user_id,
                        'Is_enquiry_mode': 'false',
                        'preferredusername': 'maker',
                        'role': 'maker',
                        'product_type': 'ELCP',
                        'userid':user_id_rob ,
                        'username':username_rob,
                        'Lc_id':'',
                        'Transaction_reference':str(ref_id),
                        'Accept':'application/json, text/plain, */*',
                        'Origin':'https://ce-tf-ui-qa-v2.psionix.dev',
                        'Referer':'https://ce-tf-ui-qa-v2.psionix.dev/'
                    }
                    time.sleep(UPLOAD_WAIT)

                    read_extraction_response = requests.get(read_extraction, headers=header004,timeout=30)
                    print(read_extraction_response.json())

                 
                    if read_extraction_response.status_code == 200:
                        # Extracting supporting document field values
                        print('supporting document extraction started for '+each['name'])
                        #supporting_document_extracted_data(read_extraction_response)
                        extracted_file_name = supporting_document_extracted_data(read_extraction_response)
    
                    else:
                        print("Failed to extract supporting documents "+ each['name'])
                    try:
                     get_goods_services(each['document_type_id'],ref_id,authorization_key,extracted_file_name)
                     #break
                    except Exception as e:
                      print('goods service api failed for '+each['document_file_name'] + ' with exception '+str(e))
                except Exception as e:
                    print("Failed to extract supporting documents " + each['name'] +'and the exception is '+str(e))

# to keep unique fields for each document types [if any fields are having multiple values,the value with high confidence
# score will be retained and other values will be removed from the Extraction sheet
def remove_duplicate_entries():
    data_file = load_workbook(filename=Excel_Path)
    act_data_sheet = data_file['Extraction_Sheet']
    for i in range(2, act_data_sheet.max_row + 1):
        file_name = act_data_sheet.cell(i, 2).value
        tag_name = act_data_sheet.cell(i, 4).value
        act_value = act_data_sheet.cell(i, 5).value
        conf_score = act_data_sheet.cell(i, 6).value
        for j in range(i + 1, act_data_sheet.max_row + 1):
            if act_data_sheet.cell(j, 2).value == file_name and act_data_sheet.cell(j, 4).value == tag_name:
                try:
                    # to check duplicates in boolean fields
                    if type(act_value) is bool:
                        print(tag_name)
                        if act_value is False and act_data_sheet.cell(j, 5).value is True:
                            act_data_sheet.delete_rows(i)
                        elif act_value is True and act_data_sheet.cell(j, 5).value is False:
                            act_data_sheet.delete_rows(j)
                        else:
                            act_data_sheet.delete_rows(j)
                    # to check duplicates in other fields
                    elif act_data_sheet.cell(j, 5).value == act_value:
                        act_data_sheet.delete_rows(j)
                    else:
                        if act_data_sheet.cell(j, 6).value == conf_score:
                            act_data_sheet.delete_rows(j)
                        elif conf_score > act_data_sheet.cell(j, 6).value:
                            act_data_sheet.delete_rows(j)
                        elif act_data_sheet.cell(j, 6).value > conf_score:
                            act_data_sheet.delete_rows(i)
                except Exception as e:
                    print("An exception occurred in row number " + str(j) + tag_name + " of the Extraction Sheet" + str(
                        e))
            elif act_data_sheet.cell(j, 2).value != file_name:
                break
        continue
    print("All the duplicate entries are removed from the Extraction sheet")
    data_file.save(filename=Excel_Path)


# to compare the actual and expected data
def compare_fields(file_name):
    # data_file = load_workbook(filename=Excel_Path)
    data_file = load_workbook(filename=file_name)
    Actual_Sheet = data_file['Extraction_Sheet']
    # to create a copy of the Expected sheet and name it as 'Results'
    sheet = data_file.copy_worksheet(data_file['Expected_Sheet'])
    sheet.title = "Results"
    Comparison_Sheet = data_file['Results']
    Comparison_Sheet.cell(1, 7).value = "Result"
    Comparison_Sheet.cell(1, 8).value = "Extracted Value"
    Comparison_Sheet.cell(1, 9).value = "Match Ratio"
    # to loop through Result sheet and extracted sheet
    for i in range(2, Comparison_Sheet.max_row + 1):
        for j in range(2, Actual_Sheet.max_row + 1):
            Expected_File_Name = Comparison_Sheet.cell(i, 3).value
            ExpectedTag = Comparison_Sheet.cell(i, 5).value
            exp_value = Comparison_Sheet.cell(i, 6).value
            File_Name = Actual_Sheet.cell(j, 2).value
            TagName = Actual_Sheet.cell(j, 4).value
            ActualValue = Actual_Sheet.cell(j, 5).value
            try:
                # check whether the expected and actual file name matches
                if str(Expected_File_Name).lower().strip() == str(File_Name).lower().strip():
                    # check whether the expected and actual tag name matches
                    if ExpectedTag not in accuracy_exclude_list:
                        if str(ExpectedTag).lower().strip() == str(TagName).lower().strip():
                            match_ratio = fuzz.token_sort_ratio(str(exp_value).lower(), str(ActualValue).lower())
                            if exp_value and ActualValue:
                                if 'amount' in str(ExpectedTag).lower() or 'weight' in str(ExpectedTag).lower() \
                                        or 'quantity' in str(ExpectedTag).lower() or 'goods' in str(ExpectedTag).lower() \
                                        or 'incoterm' in str(ExpectedTag).lower() or 'port' in str(ExpectedTag).lower():
                                    expected_match_ratio = 75
                                    fuzzy_match(ActualValue, Comparison_Sheet, i, match_ratio, expected_match_ratio)
                                elif 'address' in str(ExpectedTag).lower() or 'payment terms' in \
                                        str(ExpectedTag).lower() or 'pay to' in str(ExpectedTag).lower() \
                                        or 'declaration' in str(ExpectedTag).lower() \
                                        or 'risk clauses' in str(ExpectedTag).lower():
                                    expected_match_ratio = 85
                                    fuzzy_match(ActualValue, Comparison_Sheet, i, match_ratio, expected_match_ratio)
                                elif 'date' in str(ExpectedTag).lower():
                                    Exp_Date = parser.parse(str(exp_value)).isoformat()
                                    Act_Date = parser.parse(str(ActualValue)).isoformat()
                                    if str(Exp_Date) == str(Act_Date):
                                        Comparison_Sheet.cell(i, 7).value = 'PASS'
                                    else:
                                        Comparison_Sheet.cell(i, 7).value = 'FAIL'
                                        Comparison_Sheet.cell(i, 7).fill = redFill
                                    Comparison_Sheet.cell(i, 8).value = ActualValue
                                    Comparison_Sheet.cell(i, 9).value = match_ratio
                                else:
                                    expected_match_ratio = 90
                                    fuzzy_match(ActualValue, Comparison_Sheet, i, match_ratio, expected_match_ratio)
                            elif exp_value is None and ActualValue:
                                Comparison_Sheet.cell(i, 7).value = 'EXTRA PREDICTION'
                                Comparison_Sheet.cell(i, 8).value = ActualValue
                            break
                        elif exp_value is not None:
                            Comparison_Sheet.cell(i, 7).value = 'MISSED TO EXTRACT'
            except:
                print("an exception occurred in row number " + str(
                    i) + " value " + str(exp_value) + " please correct the value in Expected Sheet also")
    print("comparison of Actual and Expected values is completed")
    data_file.save(filename=file_name)


def fuzzy_match(ActualValue, Comparison_Sheet, i, match_ratio, expected_match_ratio):
    if match_ratio >= expected_match_ratio:
        Comparison_Sheet.cell(i, 7).value = 'PASS'
    else:
        Comparison_Sheet.cell(i, 7).value = 'FAIL'
        Comparison_Sheet.cell(i, 7).fill = redFill
    Comparison_Sheet.cell(i, 8).value = ActualValue
    Comparison_Sheet.cell(i, 9).value = match_ratio


# to calculate the accuracy for each attributes for each document type and
# to calculate the model accuracy
def accuracy(file_name):
    # Result_Sheet = pd.read_excel(Excel_Path, 'Results')
    Result_Sheet = pd.read_excel(file_name, 'Results')
    final_df = pd.DataFrame()
    final_df2 = pd.DataFrame()
    Document_Types = Result_Sheet['Document Type'].unique()
    for each_type in Document_Types:
        df = pd.DataFrame()
        df2 = pd.DataFrame()
        Doc_Type, TagName, PassCount, Expected_Count, MissedPrediction, ExtraPrediction, Accuracy = ([] for _ in
                                                                                                     range(7))
        document_type, total_expected, total_pass, Model_Accuracy = ([] for _ in range(4))
        Data = Result_Sheet.loc[Result_Sheet['Document Type'] == each_type]
        RequireData = Data[pd.notnull(Data['Expected Values'])]
        Pass_Count = RequireData.loc[RequireData['Result'] == 'PASS', 'Tag Name'].values.tolist()
        Extra_Predicted = Data.loc[Data['Result'] == 'EXTRA PREDICTION', 'Tag Name'].values.tolist()
        Missed = RequireData.loc[RequireData['Result'] == 'MISSED TO EXTRACT', 'Tag Name'].values.tolist()
        for each_tag in RequireData['Tag Name'].values:
            if each_tag.lower() not in accuracy_exclude_list:
                if each_tag not in TagName:
                    Doc_Type.append(each_type)
                    TagName.append(each_tag)
                    Expected_Count.append(RequireData['Tag Name'].values.tolist().count(each_tag))
                    PassCount.append(Pass_Count.count(each_tag))
                    MissedPrediction.append(Missed.count(each_tag))
                    ExtraPrediction.append(Extra_Predicted.count(each_tag))
                    Accuracy.append(
                        round((Pass_Count.count(each_tag) / (RequireData['Tag Name'].values.tolist().count(each_tag) +
                                                             Extra_Predicted.count(each_tag))) * 100))
        document_type.append(each_type)
        total_expected.append(sum(Expected_Count) + sum(ExtraPrediction))
        total_pass.append(sum(PassCount))
        if (sum(Expected_Count) + sum(ExtraPrediction)) != 0:
            Model_Accuracy.append(round(sum(PassCount) / (sum(Expected_Count) + sum(ExtraPrediction)) * 100))
        else:
            Model_Accuracy.append(0)

        # to store the accuracy for each attributes in each document type
        df = df.from_dict(
            {'Document Type': Doc_Type, 'Tag Name': TagName, 'Expected Count': Expected_Count,
             'Actual Pass Count': PassCount,
             'Missed to Extract': MissedPrediction, 'Extra Prediction': ExtraPrediction,
             'Accuracy': Accuracy})
        final_df = pd.concat([final_df, df])
        # to store the model accuracy
        df2 = df2.from_dict({'Models': document_type, 'Expected Count': total_expected, 'Total Pass': total_pass,
                             'Accuracy': Model_Accuracy})
        final_df2 = pd.concat([final_df2, df2])
    print(final_df)
    print(final_df2)
    time.sleep(30)
    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as wb:
        final_df.to_excel(wb, sheet_name='Attribute_Accuracy', header=True, index=False)
        final_df2.to_excel(wb, sheet_name='Model_Accuracy', header=True, index=False)

def remove_expectedsheet_entries():
    data_file = load_workbook(filename=Excel_Path)
    exp_data_sheet = data_file['Expected_Sheet']
    upload_sheet = data_file['Upload_List']
    m = None
    n = None
    for m in range(2, upload_sheet.max_row + 1):
        
        
        if  upload_sheet.cell(m, 4).value == 'No':
            deleted_count = 0
            for n in range(2, exp_data_sheet.max_row + 1):
                
                if upload_sheet.cell(m, 2).value == exp_data_sheet.cell(n-deleted_count, 3).value :
                    exp_data_sheet.delete_rows(n-deleted_count)
                    deleted_count = deleted_count + 1
        
          
        

    print("All the entries which are not required for Accuracy are removed from the Expected sheet")
    data_file.save(filename=Excel_Path)
