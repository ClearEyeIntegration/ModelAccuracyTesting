import openpyxl
import requests
import time
from openpyxl.styles import PatternFill


#ACCESS_TOKEN_API = 'https://ce-tf-api-dev.psionix.dev/jobs/access_token'    
# ACCESS_TOKEN_API = 'https://ce-tf-api-qa.psionix.dev/jobs/access_token'
#
# GOODS_API = 'https://ce-tf-api-qa.psionix.dev/jobs/goods-and-services'
# environment = 'QA'
# org_code = 'CETRQA'
# upload_api = 'https://ce-tf-api-qa.psionix.dev/jobs/files'
# bill_type_api = 'https://ce-tf-api-qa.psionix.dev/jobs/classification/result'
# list_jobs = 'https://ce-tf-api-qa.psionix.dev/jobs/presentations'
#
# list_NT_lc_txns2 = 'https://ce-tf-api-qa.psionix.dev/transactions?tag=Letter%20of%20credit&view=new&page_no=1&page_size=50'
# list_NT_lc_txns = 'https://ce-tf-api-qa.psionix.dev/transactions?tag=Letter%20of%20credit&view=new'
# doc_extraction_api = 'https://ce-tf-api-qa.psionix.dev/jobs/0/extraction/document'


#stage
ACCESS_TOKEN_API = 'https://ce-tf-api-stage.cleartrade.dev/jobs/access_token'

GOODS_API = 'https://ce-tf-api-stage.cleartrade.dev/jobs/goods-and-services'
environment = 'STAGE'
org_code = 'CETRSTAGE'
upload_api = 'https://ce-tf-api-stage.cleartrade.dev/jobs/files'
bill_type_api = 'https://ce-tf-api-stage.cleartrade.dev/jobs/classification/result'
list_jobs = 'https://ce-tf-api-stage.cleartrade.dev/jobs/presentations'

list_NT_lc_txns2 = 'https://ce-tf-api-stage.cleartrade.dev/transactions?tag=Letter%20of%20credit&view=new&page_no=1&page_size=50'
list_NT_lc_txns = 'https://ce-tf-api-stage.cleartrade.dev/transactions?tag=Letter%20of%20credit&view=new'
doc_extraction_api = 'https://ce-tf-api-stage.cleartrade.dev/jobs/0/extraction/document'



#sandbox

#ACCESS_TOKEN_API='https://sandbox-api.cleartrade.dev/jobs/access_token'
#GOODS_API = 'https://sandbox-api.cleartrade.dev/jobs/goods-and-services'
#environment = 'SANDBOX'

#org_code='JPMCSSO'


#upload_api = 'https://sandbox-api.cleartrade.dev/jobs/files'

#bill_type_api = 'https://ce-tf-api-qa.psionix.dev/jobs/classification/result'

#list_jobs = 'https://ce-tf-api-qa.psionix.dev/jobs/presentations'

#list_NT_lc_txns2 = 'https://sandbox-api.cleartrade.dev/transactions?tag=Letter%20of%20credit&view=new&page_no=1&page_size=50'
#list_NT_lc_txns = 'https://sandbox-api.cleartrade.dev/transactions?tag=Letter%20of%20credit&view=new'
#doc_extraction_api = 'https://sandbox-api.cleartrade.dev/jobs/0/extraction/document'
user_id = '47260b0a-f7b1-47b6-993c-8276e41adc84'

user_id_rob = '7fc7b4ca-d58d-4d94-b4d6-86fc05757980'
#user_id_rob = 'cd2a1333-5595-468d-b6b5-2e90528946ce'
username_rob = 'Rob Maker'
username = "maker@cleareye.ai"
password = "P@ssw0rd"
product_type = "ELCP"
branch = 'GLB'
UPLOAD_WAIT = 60
doc_type_id_charter_party = 77
doc_type_id_ci = 4
doc_type_id_cl = 44    #coverletter
doc_type_id_coo = 8
doc_type_bene_cert = 9
#qa_name = "As"              # Name/Initials of the QA resource who is running

# Give the Excel name which has upload files list and expected values
# Excel_Path = 'PL_MODELACCURACY.xlsx'
Excel_Path = 'CIMODELACCURACY.xlsx'
#Excel_Path = 'CLMODELACCURACY.xlsx'
#Excel_Path = 'PL_MODELACCURACY.xlsx'
excel_wb = openpyxl.load_workbook(Excel_Path)
excel_sheet = excel_wb['Upload_List']

# to color failed/missed/extra predicted items in Red
redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")

exclude_list = ['page_paths', 'reference_number', 'page_urls', 'file_path', 'document_type_key',
                'document_reference_id', 'document_type_id', 'document_type_name',
                'tables', 'read_results', 'document_id', 'document_file_name', 'signature_seal']

# below items present in the list will be excluded from the attribute accuracy
accuracy_exclude_list = ['drawee country', 'drawer country', 'country of beneficiary', 'country of applicant',
                         'language of document', 'country of final destination', 'country of seller',
                         'country of buyer',
                         'departure airport', 'clean bl', 'destination airport', 'exporter country',
                         'consignee country',
                         'departure date', 'seal', 'signature', 'terms and conditions', 'final destination',
                         'advance amount', 'discount amount', 'goods amount', 'goods quantity', 'goods unit price',
                         'goods gross weight', 'goods net weight', 'goods quantity', 'packages', 'goods quantity unit',
                         'goods gross weight unit', 'goods net weight unit', 'incoterm source', 'charter party',
                         'clean bl', 'transhipment',
                         'right to tranship', 'freight forwarder', 'split charges', 'onboard notation',
                         'clean airway bill',
                         'signed as carrier or agent', 'issued for consignor or shipper', 'endorsement',
                         'warehouse clause present',
                         'risk cover match', 'all risks covered', 'issued for consignor or shipper']


# To get the access token
def get_access_token():
    try:
        header = {
            "username": username,
            "password": password,
            "organization_code":org_code

        }
        response = requests.get(ACCESS_TOKEN_API, headers=header,timeout=30)
        #print(response.text)
        time.sleep(3)
        return response

    except:
        print('Couldnt get the Access Token')


def assign_lc_to_user(nt_docv_id,nt_refid, authorization_key,product_type,ref_id,wid):
    # try:
   
    assign_api = 'https://ce-tf-api-qa.psionix.dev/transactions/' + str(nt_refid) + '/assign'
    #assign_api = 'https://sandbox-auth.cleartrade.dev/transactions/' + str(nt_refid) + '/assign'
    print(assign_api)

    headers = {
        'organization_code': org_code,
        'Authorization': authorization_key,
        'is_enquiry_mode': 'false',
        'preferredusername': 'maker',
        'role': 'maker',
        'product_type': product_type,
        'transaction_reference':nt_refid,
        'userid':user_id_rob ,
        'username':username_rob

    }
    
   

    payload = {"product_type":product_type,
                "assign_to":user_id_rob,
                "assigned_by":user_id_rob,
                "assigned_to_name":user_id_rob,
                "assigned_by_name":user_id_rob,
                "assigned_role":"maker",
                "work_item_ids":[nt_docv_id]
                }
    #assign_response = requests.post(assign_api, headers=headers, json=payload) #amitabh
    assign_response = requests.put(assign_api, headers=headers, json=payload)
    if assign_response.status_code == 200:
        print("Successfully assigned " + ref_id + " WorkItem: " + wid)
    else:
        print(assign_response.status_code)
    #print(assign_response)


    return assign_response

